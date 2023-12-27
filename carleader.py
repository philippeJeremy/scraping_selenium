import os
import time
import psycopg2
import pandas as pd
from datetime import datetime
from selenium import webdriver
from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

dimension = pd.read_excel('Copie de Ranking.xlsx', sheet_name='Feuil1')
liste_articles = dimension['Dimension'].to_list()
# liste_articles = ['2055516V91']

saisons = ["été", "4 saisons"]

marques = ["Michelin", "Continental", "Bridgestone",  'Dunlop',  'Goodyear',
           'Kleber', 'Firestone', 'Vredestein', "Hankook", 'Uniroyal', 'Kormoran', 'Tourador']

load_dotenv()


def select_saison(driver, saison):
    """
    fonction de selection de saison
    """
    saison_ete = driver.find_element(By.ID, "customCheck2")
    if 'été' in saison:
        if saison_ete.is_selected():
            pass
        else:
            driver.execute_script("arguments[0].click();", saison_ete)
            time.sleep(1)
    else:
        if saison_ete.is_selected():
            driver.execute_script("arguments[0].click();", saison_ete)
            time.sleep(1)
    saison_4saison = driver.find_element(By.ID, "customCheck4")
    time.sleep(1)
    if '4 saisons' in saison:
        if saison_4saison.is_selected():
            pass
        else:
            driver.execute_script(
                "arguments[0].click();", saison_4saison)
            time.sleep(1)
    else:
        if saison_4saison.is_selected():
            driver.execute_script(
                "arguments[0].click();", saison_4saison)
            time.sleep(1)
    saison_hiver = driver.find_element(By.ID, "customCheck3")
    time.sleep(1)
    if 'hiver' in saison:
        if saison_hiver.is_selected():
            pass
        else:
            driver.execute_script(
                "arguments[0].click();", saison_hiver)
            time.sleep(1)
    else:
        if saison_hiver.is_selected():
            driver.execute_script(
                "arguments[0].click();", saison_hiver)
            time.sleep(1)


def login(driver):
    """
    fonction de login du site
    """
    login = driver.find_element(By.ID, "login_form_customer_code")
    login.send_keys(os.getenv("CARLEADER_LOG"))
    time.sleep(2)
    password = driver.find_element(By.ID, "login_form_password")
    password.send_keys(os.getenv("CARLEADER_PASS"))
    time.sleep(2)
    password.send_keys(Keys.ENTER)
    time.sleep(15)


def search_article(driver, marque, article):
    """
    fonction de recherche d'article
    """
    try:
        time.sleep(1)
        search_article = driver.find_element(By.ID, "validationServer01")
        driver.execute_script("arguments[0].value = '';", search_article)
        search_article.clear()
        search_article.send_keys(
            f"{article[0:3]+'/'+ f'{article[3:5]}' + 'R' + f'{article[5:7]}'}")
        time.sleep(1)
        if len(article) <= 10:
            charge = Select(driver.find_element(By.ID, "validationServer02"))
            charge.select_by_value(f'{article[8:10]}')
        if len(article) > 10:
            charge = Select(driver.find_element(By.ID, "validationServer02"))
            charge.select_by_value(f'{article[8:11]}')
        time.sleep(1)
        vitesse = Select(driver.find_element(By.ID, "validationServer03"))
        vitesse.select_by_value(f'{article[7]}')
        time.sleep(1)
        search_article.send_keys(Keys.ENTER)
        time.sleep(30)
    except Exception as ex:
        print(ex)


def extract_article_info(driver, article, liste_article):
    date_du_jour = datetime.now().strftime("%d-%m-%Y")
    ind_C = article.find_element(
        By.XPATH, 'div[2]/div/div/div[2]/div/div/ul/li[1]')
    tab_c = ind_C.text.split()
    ind_V = article.find_element(
        By.XPATH, 'div[2]/div/div/div[2]/div/div/ul/li[2]')
    tab_v = ind_V.text.split()
    code = article.find_element(
        By.XPATH, 'div/div/div[1]/div/div[2]/div/div[1]/span')
    text_code = code.text.replace("/", "").replace("-", "").replace(" ", "")
    marque = article.find_element(
        By.XPATH, 'div/div/div[1]/div/div[2]/div/div[1]/p/strong')
    designation = article.find_element(
        By.XPATH, 'div/div/div[1]/div/div[2]/div/div[1]/span')
    prix = article.find_element(
        By.XPATH, 'div[1]/div/div[4]/div/div[2]/div/div[2]/span[2]')
    saison = article.find_element(
        By.XPATH, 'div[2]/div/div/div[2]/div/div/ul/li[3]')
    saison_tab = saison.text.split()
    if saison_tab[-2] == '4':
        texte_saison = saison_tab[-2] + ' ' + saison_tab[-1]
    else:
        texte_saison = saison_tab[-1]

    article_info = {
        'Code': text_code[0:7] + tab_v[-1] + tab_c[-1],
        'Marque': marque.text.replace(' ', ''),
        'Code article': designation.text.replace(" ", "").replace("-", "").replace("/", "").replace(".", ""),
        'Designation': designation.text,
        'Prix': prix.text.replace('€', ''),
        'Saison': texte_saison,
        'Date': date_du_jour,
        'Site': 'Carleader'
    }

    if liste_article != article_info["Code"]:
        return None
    else:
        print(article_info)
        return article_info


def save_donnees_excel(data, marques):
    df = pd.DataFrame(data)
    df.replace("", pd.NA, inplace=True)
    df.dropna(inplace=True)

    workbook = load_workbook('Scrap_site_web.xlsx')
    with pd.ExcelWriter('Scrap_site_web.xlsx', if_sheet_exists="overlay", mode='a', engine='openpyxl') as writer:
        for marque in marques:

            df_marque = df[df['Marque'] == marque]

            df_marque.to_excel(writer, sheet_name=marque,
                               startrow=workbook[marque].max_row, index=False, header=False)


def save_donnees_sql(data):
    HOST = os.getenv('HOST')
    PORT = os.getenv('PORT')
    ID = os.getenv('ID')
    DATABASE = os.getenv('DATABASE')
    PASSWORD = os.getenv('PASSWORD')

    connection_string = f"postgresql+psycopg2://{ID}:{PASSWORD}@{HOST}:{PORT}/{DATABASE}"
    engine = create_engine(connection_string)

    data = [item for item in data if item is not None]

    df = pd.DataFrame(data)
    df.replace("", pd.NA, inplace=True)
    df.dropna(inplace=True)

    df.to_sql('scrap_pneu', engine, if_exists='append', index=False)


def carleader_scrap(liste_articles, saison_select, marque_spe):

    # connexion au site web
    driver = webdriver.Firefox('')
    driver.get("https://b2b.carleader.com/")
    time.sleep(2)

    login(driver)

    data = []

    for marque in marque_spe:

        time.sleep(1)
        select_marque = driver.find_element(
            By.CLASS_NAME, 'multiselect__input')
        select_marque.send_keys(marque)
        time.sleep(1)
        select_marque.send_keys(Keys.ENTER)
        time.sleep(1)
        select_marque.send_keys(Keys.ESCAPE)

    for liste_article in liste_articles:
        select_saison(driver, saison_select)
        time.sleep(1)
        search_article(driver, marque, liste_article)
        driver.find_element(
            By.XPATH, '/html/body/main/div/div/div[2]/div/div[2]/div[1]/div/div[2]/form/div[3]/label').click()
        time.sleep(2)
        articles = driver.find_elements(
            By.CLASS_NAME, 'row.mx-0.border-top.border-gray.is-tertiary')

        for article in articles:
            article_info = extract_article_info(
                driver, article, liste_article)
            if article_info:
                data.append(article_info)

    # save_donnees_excel(data, marques)
    save_donnees_sql(data)

    driver.quit()


if __name__ == "__main__":
    carleader_scrap(liste_articles, saisons, marques)
