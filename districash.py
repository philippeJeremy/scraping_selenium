import os
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select

# dimension = pd.read_excel('Copie de Ranking.xlsx', sheet_name='Feuil1')
# liste_articles = dimension['Dimension'].to_list()
liste_articles = ['2056516T107']

saisons_disti = ["été", "hiver", "4 saisons"]
# marques = ["MICH", "CONT", "BRID",  'DUNL',  'GOAR', 'TOUR',
#            'KLEB', 'FIRE', 'VRED', "HANK", 'UNIR', 'KORM']

marques = ["Michelin", "Continental", "Bridgestone",  'Dunlop',  'Goodyear',
           'Kleber', 'Firestone', 'Vredestein', "Hankook", 'Uniroyal', 'Kormoran', 'Tourador']

load_dotenv()


def login(driver):
    login_elem = driver.find_element(By.NAME, "login")
    login_elem.send_keys(os.getenv('DISTRICASH_LOG'))

    password_elem = driver.find_element(By.NAME, "password")
    password_elem.send_keys(os.getenv('DISTRICASH_PASS'))
    password_elem.send_keys(Keys.ENTER)


def select_saisons(driver):
    try:
        saison_ete = driver.find_element(By.NAME, "saisonETE")
        saison_4saison = driver.find_element(By.NAME, "saisonETEHIVER")
        saison_hiver = driver.find_element(By.NAME, "saisonHIVER")

        if not saison_ete.is_selected():
            driver.execute_script("arguments[0].click();", saison_ete)
            time.sleep(1)

        if not saison_4saison.is_selected():
            driver.execute_script("arguments[0].click();", saison_4saison)
            time.sleep(1)

        if not saison_hiver.is_selected():
            driver.execute_script("arguments[0].click();", saison_hiver)
            time.sleep(1)

    except Exception as ex:
        print(ex)


def select_saisons_2(driver, saison_select):
    try:
        saison_ete = driver.find_element(By.ID, "ete_afficher")
        saison_4saison = driver.find_element(By.ID, "hiver_afficher")
        saison_hiver = driver.find_element(By.ID, "ete_hiver")

        if 'été' in saison_select:
            if not saison_ete.is_selected():
                driver.execute_script("arguments[0].click();", saison_ete)
                time.sleep(1)
        else:
            if saison_ete.is_selected():
                driver.execute_script("arguments[0].click();", saison_ete)
                time.sleep(1)

        if '4 saisons' in saison_select:
            if not saison_4saison.is_selected():
                driver.execute_script("arguments[0].click();", saison_4saison)
                time.sleep(1)
        else:
            if saison_4saison.is_selected():
                driver.execute_script("arguments[0].click();", saison_4saison)
                time.sleep(1)

        if 'hiver' in saison_select:
            if not saison_hiver.is_selected():
                driver.execute_script("arguments[0].click();", saison_hiver)
                time.sleep(1)
        else:
            if saison_hiver.is_selected():
                driver.execute_script("arguments[0].click();", saison_hiver)
                time.sleep(1)

    except Exception as ex:
        print(ex)


def extract_article_info(article, driver, liste_article):
    date_du_jour = datetime.now().strftime("%d-%m-%Y")
    try:
        marque = article.find_element(
            By.XPATH, 'td[9]/img').get_attribute('title')
        code = article.find_element(By.XPATH, 'td[10]')
        designation = article.find_element(By.XPATH, 'td[11]/a')
        charge = Select(driver.find_element(By.ID, "liste_indcharge"))
        ind_C = charge.first_selected_option
        valeur_ind_C = ind_C.get_attribute("value")
        prix = article.find_element(By.XPATH, 'td[16]/span/b')
        saison = article.find_element(
            By.XPATH, 'td[29]/b/img').get_attribute('title')
        design = designation.text.split(' ')
        digits = ''
        letters = ''
        last = ''
        for char in design[3]:
            if char.isdigit():
                digits += char
            elif char.isalpha():
                letters += char

        for i in design[4:]:
            last += i
        num = design[2]
        code_article = num[0:9].replace('/', '') + letters + digits + last

        article_info = {
            'Code': code.text[0:8] + valeur_ind_C,
            'Marque': marque.capitalize(),
            'Code article': code_article,
            'Prix': prix.text,
            'Saison': saison,
            'Date': date_du_jour,
            'Site': 'Districash'
        }

        if liste_article != article_info["Code"]:
            return None
        else:
            print(article_info)
            return article_info
    except Exception as ex:
        print(ex)
        return None


def search_article(driver, liste_article):
    while True:
        try:
            search_article = driver.find_element(By.ID, "code_article_input")
            search_article.clear()
            search_article.send_keys(f"{liste_article[0:8]}")
            time.sleep(2)
            # ajout indice de charge exemple'91'
            if len(liste_article) <= 10:
                charge = Select(driver.find_element(By.ID, "liste_indcharge"))
                charge.select_by_value(f'{liste_article[8:10]}')
                time.sleep(2)
            if len(liste_article) > 10:
                charge = Select(driver.find_element(By.ID, "liste_indcharge"))
                charge.select_by_value(f'{liste_article[8:11]}')
                time.sleep(2)
            driver.find_element(By.ID, 'envoie_formulaire').click()
            time.sleep(80)
            driver.find_element(By.XPATH, '//*[@id="loader"]/div/p/a').click()

        except Exception as ex:
            break


def save_donnees_excel(data, marques):
    df = pd.DataFrame(data)
    df.replace("", pd.NA, inplace=True)
    df.dropna(inplace=True)
    df = df[df['saison'] != 'hiver']
    df.replace("été / hiver", '4 Saisons', inplace=True)
    df.replace("été", 'Eté', inplace=True)

    workbook = load_workbook('Scrap_site_web.xlsx')
    with pd.ExcelWriter('Scrap_site_web.xlsx', if_sheet_exists="overlay", mode='a', engine='openpyxl') as writer:
        for marque in marques:

            df_marque = df[df['Marque'] == marque]

            df_marque.to_excel(writer, sheet_name=marque,
                               startrow=workbook[marque].max_row, index=False, header=False)


def save_donnees_sql(data, marques):
    HOST = os.getenv('HOST')
    PORT = os.getenv('PORT')
    ID = os.getenv('ID')
    DATABASE = os.getenv('DATABASE')
    PASSWORD = os.getenv('PASSWORD')

    # data = [item for item in data if item is not None]

    df = pd.DataFrame(data)
    # df.replace("", pd.NA, inplace=True)
    # df.dropna(inplace=True)
    # print(df.head(5))
    df = df[df['Marque'].isin(marques)]
    df = df[df['Saison'] != 'hiver']
    df.replace("été / hiver", '4 Saisons', inplace=True)
    df.replace("été", 'Eté', inplace=True)

    connection_string = f"postgresql+psycopg2://{ID}:{PASSWORD}@{HOST}:{PORT}/{DATABASE}"
    engine = create_engine(connection_string)

    df.to_sql('scrap_pneu', engine, if_exists='append', index=False)


def districash_scrap(liste_articles, saisons, marques):
    # connexion au site
    with webdriver.Firefox('') as driver:
        driver.get(
            "https://districash.inoshop.net/connect_v2/index.php")

        # identification
        time.sleep(2)
        login(driver)
        time.sleep(7)

        # tableau de récuperation de données
        data = []

        for liste_article in liste_articles:
            driver.refresh()
            time.sleep(10)
            select_nb_element = driver.find_element(
                By.ID, 'liste_top_article')
            select_nb = Select(select_nb_element)
            time.sleep(1)
            select_nb.select_by_value("80")
            time.sleep(1)
            select_saisons(driver)
            time.sleep(1)
            search_article(driver, liste_article)

            articles = driver.find_elements(
                By.XPATH, "//*[@id='TABLEAU_RECHERCHE_ARTICLE']/tbody/tr")
            # recuperation des info
            for article in articles:
                article_info = extract_article_info(
                    article, driver, liste_article)
                if article_info:
                    data.append(article_info)

    # récuperation des données avec pandas mise en forem et sauvegarde
        # save_donnees_excel(data, marques)

        save_donnees_sql(data, marques)

    driver.quit()


if __name__ == "__main__":
    districash_scrap(liste_articles, saisons_disti, marques)
