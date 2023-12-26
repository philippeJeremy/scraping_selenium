import os
import time
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC

# dimension = pd.read_excel('Copie de Ranking.xlsx', sheet_name='Feuil1')
# liste_articles = dimension['Dimension'].to_list()
liste_articles = ['2055516V91']

saisons = ["été", "4 saisons"]

marques = ["Michelin", "Continental", "Bridgestone",  'Dunlop',  'Goodyear',
           'Kleber', 'Firestone', 'Vredestein', "Hankook", 'Uniroyal', 'Kormoran', 'Tourador']

load_dotenv()


def login(driver):
    username_input = driver.find_element(By.ID, "nx-login-form-username")
    username_input.send_keys(os.getenv('GETTYGO_LOG'))

    password_input = driver.find_element(By.ID, "nx-login-form-password")
    password_input.send_keys(os.getenv('GETTYGO_PASS'))
    password_input.send_keys(Keys.ENTER)


def wait_for_element(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


def select_season(driver, season_select):
    if 'été' in season_select:
        wait_for_element(driver, By.XPATH, '//*[@title="Été"]').click()
        time.sleep(1)
    if '4 saisons' in season_select:
        wait_for_element(driver, By.XPATH, '//*[@title="4 saisons"]').click()
        time.sleep(1)
    if 'hiver' in season_select:
        wait_for_element(driver, By.XPATH, '//*[@title="Hiver"]').click()
        time.sleep(1)


def extract_article_info(article, driver, liste_article):
    try:
        date_du_jour = datetime.now().strftime("%d-%m-%Y")
        dim = article.find_element(
            By.XPATH, 'div[1]/div/div[2]/div/div[1]/div')
        marque = article.find_element(
            By.XPATH, 'div[1]/div/div[1]/div/div[2]/div[1]/b')
        prix = article.find_element(
            By.XPATH, 'div[3]/div/div[2]/div/div[2]/div[1]/div[2]/div')
        saison = article.find_element(
            By.XPATH, 'div[1]/div/div[1]/div/div[2]/div[3]')
        indice = article.find_element(
            By.XPATH, 'div[1]/div/div[2]/div/div[2]')
        profil = article.find_element(
            By.XPATH, "div[2]/div/div[2]/span/b")
        code = dim.text.replace('/', '').replace('R', '')
        ind_split = indice.text.split()
        indice_V = ind_split[1]
        indice_C = ind_split[0]

        article_info = {
            'Code': code.replace(' ', '') + indice_V + indice_C,
            'Marque': marque.text.capitalize(),
            'Code article': code.replace(' ', '') + 'TL' + indice_V + indice_C + profil.text.replace(' ', ""),
            'Designation': dim.text + 'TL' + indice_V + indice_C + profil.text,
            'Prix': prix.text.replace('€', '').replace(',', '.'),
            'Saison': saison.text.replace('VL', '').replace(' ', ''),
            'Date': date_du_jour,
            'Site': 'Gettygo'
        }

        if liste_article != article_info["Code"]:
            return None
        else:
            print(article_info)
            return article_info
    except Exception as ex:
        print(ex)
        return None


def search_article(driver, marques, liste_article):
    try:
        time.sleep(1)
        reset = driver.find_element(
            By.XPATH, '//*[@id="nx-main-background-container"]/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-quick-options/div/div[2]/div[1]/div[3]/nx-button[2]/div/button/span')
        reset.click()
        time.sleep(1)
        vl = driver.find_element(
            By.XPATH, "//*[@id='nx-main-background-container']/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-quick-options/div/div[1]/div[2]/label[1]")
        driver.execute_script("arguments[0].click();", vl)
        select_season(driver, saisons)
        driver.find_element(
            By.XPATH, '//*[@id="nx-main-background-container"]/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-quick-options/div/div[2]/div[1]/div[2]/nx-brand-selection/div').click()
        time.sleep(1)
        search_marque = driver.find_element(
            By.XPATH, '//*[@id="nx-main-background-container"]/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-quick-options/div/div[2]/div[1]/div[2]/nx-brand-selection/div/div[1]/input')
        time.sleep(1)
        for i in range(len(marques)):
            time.sleep(1)
            search_marque.send_keys(marques[i])
            time.sleep(1)
            search_marque.send_keys(Keys.ENTER)
            time.sleep(1)
            search_article = driver.find_element(
                By.ID, "nx-tyre-quick-options-matchcode")
            search_article.clear()
            time.sleep(1)

        search_article.send_keys(f"{liste_article[0:8]}")
        time.sleep(2)
        search_article.send_keys(Keys.ENTER)

        time.sleep(20)
    except Exception as ex:
        print(ex)


def save_donnees_excel(data, marques):
    df = pd.DataFrame(data)

    workbook = load_workbook('Scrap_site_web.xlsx')
    with pd.ExcelWriter('Scrap_site_web.xlsx', if_sheet_exists="overlay", mode='a', engine='openpyxl') as writer:
        for marque in marques:

            df_marque = df[df['Marque'] == marque]

            df_marque.to_excel(writer, sheet_name=marque,
                               startrow=workbook[marque].max_row, index=False, header=False)


def save_donnees_sql(data):
    HOST = os.getenv('HOST')
    PORT = os.getenv('PORT')
    ID = os.getenv('ID')
    DATABASE = os.getenv('DATABASE')
    PASSWORD = os.getenv('PASSWORD')

    connection_string = f"postgresql+psycopg2://{ID}:{PASSWORD}@{HOST}:{PORT}/{DATABASE}"
    engine = create_engine(connection_string)

    data = [item for item in data if item is not None]

    df = pd.DataFrame(data)
    df.replace('ÉTÉ', 'Eté', inplace=True)
    df.replace('4SAISONS', '4 Saisons', inplace=True)

    df.to_sql('scrap_pneu', engine, if_exists='append', index=False)


def gettygo_scrap(liste_articles, saisons, marques):
    driver = webdriver.Firefox('')

    driver.get(
        "https://www.gettygo.fr")
    time.sleep(2)

    login(driver)
    time.sleep(15)

    data = []

    for liste_article in liste_articles:
        try:
            search_article(driver, marques, liste_article)

            driver.find_element(
                By.XPATH, '/html/body/div[1]/div/nx-landing-page/div/nx-main/div/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-search-panel/div/div[2]/div[1]/div[2]/div[2]/nx-page-size-selector/div/div/div[1]/span/i').click()
            time.sleep(2)
            driver.find_element(
                By.XPATH, '/html/body/div[1]/div/nx-landing-page/div/nx-main/div/div/div/div/section/div/nx-tyre-quick-search/div/nx-tyre-search-panel/div/div[2]/div[1]/div[2]/div[2]/nx-page-size-selector/div/div/ul/li/div[6]/a').click()
            time.sleep(5)
            articles = driver.find_elements(
                By.CLASS_NAME, "row.nx-table-body.nx-table-body-no-alternating-lines")
            for article in articles:
                try:
                    article_info = extract_article_info(
                        article, driver, liste_article)
                    data.append(article_info)
                except Exception as ex:
                    print(ex)
                    pass
        except Exception as ex:
            print(ex)
            break

    save_donnees_sql(data)

    driver.quit()


if __name__ == "__main__":
    gettygo_scrap(liste_articles, saisons, marques)
